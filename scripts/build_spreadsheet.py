#!/usr/bin/env python3
from __future__ import annotations

import csv
import os
import re
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
IN_ACTIVITY = os.path.join(ROOT, "Encore_Combined_Report.csv")
IN_SUMMARY = os.path.join(ROOT, "Encore_Summary.csv")
OUT_XLSX = os.path.join(ROOT, "Encore_Report.xlsx")


def read_csv(path: str) -> Tuple[List[str], List[List[str]]]:
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return [], []
    raw_header = rows[0]
    data = rows[1:]
    max_cols = max([len(raw_header)] + [len(r) for r in data] + [0])

    # Pad header to match widest row (Excel tables require a header for every column)
    if len(raw_header) < max_cols:
        raw_header = raw_header + [""] * (max_cols - len(raw_header))

    header: List[str] = []
    used = set()
    for i, h in enumerate(raw_header, start=1):
        name = (h or "").strip()
        if not name:
            name = f"Column{i}"
        # Ensure uniqueness (Excel table headers must be unique strings)
        base = name
        n = 2
        while name in used:
            name = f"{base}_{n}"
            n += 1
        used.add(name)
        header.append(name)
    # Pad rows so every row has the same column count (avoid shifting in Excel)
    padded: List[List[str]] = []
    for r in data:
        if len(r) < max_cols:
            r = r + [""] * (max_cols - len(r))
        padded.append(r)
    return header, padded


def autosize_columns(ws, max_rows: int = 2000, max_width: int = 60) -> None:
    widths = {}
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx > max_rows:
            break
        for c_idx, v in enumerate(row, start=1):
            s = "" if v is None else str(v)
            widths[c_idx] = min(max(widths.get(c_idx, 0), len(s) + 2), max_width)
    for c_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(c_idx)].width = max(10, w)


def add_table(ws, name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def sanitize_sheet_title(title: str, used: set) -> str:
    # Excel: max 31 chars; cannot contain : \ / ? * [ ]
    t = (title or "").strip() or "Sheet"
    t = re.sub(r"[:\\\\/?*\\[\\]]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    t = t[:31].rstrip()
    if not t:
        t = "Sheet"
    base = t
    i = 2
    while t in used:
        suffix = f" {i}"
        t = (base[: 31 - len(suffix)].rstrip() + suffix)[:31]
        i += 1
    used.add(t)
    return t


def sanitize_table_name(name: str, used: set) -> str:
    # Excel table names: must start with letter/_; only letters/numbers/underscore; no spaces
    n = re.sub(r"[^A-Za-z0-9_]", "_", (name or "").strip())
    if not n or not re.match(r"^[A-Za-z_]", n):
        n = f"T_{n}" if n else "T_Table"
    base = n
    i = 2
    while n in used:
        n = f"{base}_{i}"
        i += 1
    used.add(n)
    return n


def write_sheet(wb: Workbook, title: str, header: List[str], rows: List[List[str]], table_name: str) -> None:
    ws = wb.create_sheet(title)
    ws.append(header)
    for r in rows:
        ws.append(r)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Header styling
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    autosize_columns(ws)
    add_table(ws, table_name)


def main() -> int:
    if not os.path.exists(IN_ACTIVITY):
        raise SystemExit(f"Missing: {IN_ACTIVITY}")
    if not os.path.exists(IN_SUMMARY):
        raise SystemExit(f"Missing: {IN_SUMMARY}")

    act_h, act_rows = read_csv(IN_ACTIVITY)
    sum_h, sum_rows = read_csv(IN_SUMMARY)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    used_sheets = set()
    used_tables = set()

    # Always keep a full activity sheet
    write_sheet(
        wb,
        sanitize_sheet_title("Activity_All", used_sheets),
        act_h,
        act_rows,
        sanitize_table_name("ActivityAll", used_tables),
    )

    # Split activity into separate tabs by Data Type
    try:
        dt_idx = act_h.index("Data Type")
    except ValueError:
        dt_idx = -1

    if dt_idx >= 0:
        groups: Dict[str, List[List[str]]] = {}
        for r in act_rows:
            dtype = (r[dt_idx] if dt_idx < len(r) else "") or "Unknown"
            dtype = str(dtype).strip() or "Unknown"
            groups.setdefault(dtype, []).append(r)

        # Sort: bigger groups first
        for dtype, rows in sorted(groups.items(), key=lambda kv: (-len(kv[1]), kv[0].lower())):
            sheet_title = sanitize_sheet_title(dtype, used_sheets)
            table_name = sanitize_table_name(f"Type_{dtype}", used_tables)
            write_sheet(wb, sheet_title, act_h, rows, table_name)

    write_sheet(
        wb,
        sanitize_sheet_title("Summary", used_sheets),
        sum_h,
        sum_rows,
        sanitize_table_name("Summary", used_tables),
    )

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

