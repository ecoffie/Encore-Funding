#!/usr/bin/env python3
"""Rebuild report-data.js from the live Google Sheets XLSX export.

Usage:
    python3 scripts/build_from_live.py

It downloads the latest XLSX from the shared Google Sheet, parses every
typed sheet (Sponsored Post, Sponsored Video, Webinar, Post, Article,
Sponsored Episode, Custom Video, Ad Spot), deduplicates, and writes a
fresh report-data.js.

Set SHEETS_ID env var to override the default spreadsheet ID.
"""
import csv, datetime, io, json, os, re, sys, urllib.request

ROOT   = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
JS_OUT = os.path.join(ROOT, "report-data.js")

SHEETS_ID = os.environ.get(
    "SHEETS_ID",
    "1gjZ5Hv6Kxy9nrxxtyyg4EdqAHBmaSs8r3h07pz2i4PY",
)
XLSX_URL = f"https://docs.google.com/spreadsheets/d/{SHEETS_ID}/export?format=xlsx"

TYPED_SHEETS = [
    "Sponsored Post", "Sponsored Video", "Webinar",
    "Custom Video", "Post", "Ad Spot", "Article", "Sponsored Episode",
]

# Channels that store their episode link in the Attendees column
PODCAST_CHANNELS = {"Podcast"}


def _import_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        sys.exit("openpyxl is required: pip install openpyxl")


def to_month_key(v) -> str:
    if isinstance(v, (datetime.datetime, datetime.date)):
        return f"{v.year}-{v.month:02d}"
    if isinstance(v, str):
        s = v.strip().lower()
        MONTHS = {
            "january": 1, "february": 2, "march": 3, "april": 4,
            "may": 5, "june": 6, "july": 7, "august": 8,
            "september": 9, "october": 10, "november": 11, "december": 12,
        }
        for name, num in MONTHS.items():
            if s.startswith(name):
                m = re.search(r"(\d{4})", s)
                yr = int(m.group(1)) if m else 2025
                return f"{yr}-{num:02d}"
    return None


def to_iso_date(v) -> str:
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        v = v.strip()
        for fmt in ("%d %b %Y", "%d %B %Y", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.datetime.strptime(v, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
    return None


def to_int(v) -> int:
    try:
        f = float(str(v).strip().replace(",", ""))
        return int(round(f))
    except (TypeError, ValueError):
        return None


def is_url(v) -> bool:
    return bool(v) and str(v).strip().startswith("http")


def download_xlsx(url: str) -> bytes:
    print(f"Downloading: {url}")
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as r:
        data = r.read()
    print(f"  Downloaded {len(data):,} bytes")
    return data


def parse_sheets(xlsx_bytes: bytes) -> list:
    openpyxl = _import_openpyxl()
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    items = []
    seen: set[tuple] = set()   # (month, channel, title, date) dedup key

    for sheet_name in TYPED_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        for row in rows[1:]:
            # Skip completely empty rows
            if not any(x is not None for x in row):
                continue

            mk = to_month_key(row[0])
            if not mk:
                continue

            channel   = str(row[1] or "").strip()
            data_type = str(row[2] or "").strip()
            date_v    = to_iso_date(row[3])
            title     = str(row[4] or "").strip()
            link_raw  = str(row[9] or "").strip() if row[9] is not None else ""
            att_raw   = str(row[10] or "").strip() if len(row) > 10 and row[10] is not None else ""

            # Podcast links live in the Attendees column
            link = link_raw if is_url(link_raw) else None
            attendees_val = None
            if channel in PODCAST_CHANNELS and is_url(att_raw):
                link = att_raw
            elif is_url(att_raw):
                pass  # keep attendees as numeric below
            else:
                attendees_val = to_int(att_raw) if att_raw else None

            # Dedup: same month + channel + title (ignore duplicate URL variants)
            # Keep the row with the higher impressions/views when there are dupes
            dup_key = (mk, channel, title, date_v)
            if dup_key in seen:
                continue
            seen.add(dup_key)

            item = {
                "month":         mk,
                "channel":       channel,
                "type":          data_type,
                "title":         title,
                "date":          date_v,
                "link":          link,
                "impressions":   to_int(row[5]),
                "views":         to_int(row[6]),
                "engagements":   to_int(row[7]),
                "clicks":        to_int(row[8]),
                "attendees":     attendees_val,
                "chat_count":    to_int(row[11]) if len(row) > 11 else None,
                "email_sends":   to_int(row[12]) if len(row) > 12 else None,
                "article_views": to_int(row[13]) if len(row) > 13 else None,
                "interactions":  to_int(row[14]) if len(row) > 14 else None,
            }
            items.append(item)

    items.sort(key=lambda x: (x["month"], x["date"] or "", x["title"]))
    return items


def build_aggregates(items: list) -> dict:
    from collections import defaultdict
    agg: dict = {}
    by_month: dict = defaultdict(list)
    for it in items:
        by_month[it["month"]].append(it)

    def s(lst, key):
        return sum(it.get(key) or 0 for it in lst)

    for mk, lst in by_month.items():
        totals = {
            "impressions":   s(lst, "impressions"),
            "views":         s(lst, "views"),
            "engagements":   s(lst, "engagements"),
            "clicks":        s(lst, "clicks"),
            "attendees":     s(lst, "attendees"),
            "chat_count":    s(lst, "chat_count"),
            "email_sends":   s(lst, "email_sends"),
            "article_views": s(lst, "article_views"),
            "items":         len(lst),
        }
        by_channel: dict = {}
        for it in lst:
            ch = it["channel"]
            if ch not in by_channel:
                by_channel[ch] = {k: 0 for k in totals}
            for k in totals:
                by_channel[ch][k] = by_channel[ch].get(k, 0) + (it.get(k) or 0)
        agg[mk] = {"totals": totals, "by_channel": by_channel}
    return agg


def build_series(months: list, agg: dict) -> dict:
    keys = ["impressions", "views", "engagements", "clicks",
            "items", "attendees", "email_sends", "article_views"]
    return {k: [(agg.get(m, {}).get("totals", {}).get(k) or 0) for m in months] for k in keys}


def main():
    xlsx = download_xlsx(XLSX_URL)
    items = parse_sheets(xlsx)

    months = sorted(set(it["month"] for it in items))
    agg    = build_aggregates(items)
    series = build_series(months, agg)

    out = {
        "generated_at": datetime.datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        "source":       f"https://docs.google.com/spreadsheets/d/{SHEETS_ID}",
        "months":       months,
        "items":        items,
        "aggregates":   agg,
        "series":       series,
    }

    js = "window.REPORT_DATA = " + json.dumps(out, ensure_ascii=False, separators=(",", ":")) + ";\n"
    with open(JS_OUT, "w", encoding="utf-8") as f:
        f.write(js)

    print(f"\n✓ Wrote {JS_OUT}")
    print(f"  {len(months)} months · {len(items)} items")
    print("\nItems per month:")
    for mk in months:
        t = agg[mk]["totals"]
        chs = list(agg[mk]["by_channel"].keys())
        print(f"  {mk}: {t['items']:2d} items  imp={t['impressions']:>7,}  eng={t['engagements']:>5,}  channels={chs}")


if __name__ == "__main__":
    main()
